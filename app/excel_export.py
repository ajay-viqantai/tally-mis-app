"""
Builds an Excel workbook for one company+year.

Ported from legacy/tally_mis_report_generic.py's write_report_to_excel()
-- same sheet-per-list, bold-header, auto-width logic. Pulls its data
from cache.py instead of a fresh run(), so the numbers always match
what the drill-down pages already show.
"""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

from app import cache
from app.exports import EXPORTS_DIR


def _write_sheet(wb: openpyxl.Workbook, sheet_name: str, rows: list[dict]) -> None:
    ws = wb.create_sheet(title=sheet_name[:31])
    columns = list(rows[0].keys()) if rows else ["Particulars"]
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in rows:
        ws.append([r.get(c, "") for c in columns])
    for col_cells in ws.columns:
        width = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(width + 2, 10), 45)


def build_workbook(company: str, from_date: str, to_date: str) -> Path:
    data = cache.get_annual_data(company, from_date, to_date)
    report = data["report"]

    monthly_pl = cache.get_monthly_pl(company, from_date, to_date)
    cashflow = cache.get_cashflow(company, from_date, to_date)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name, rows in report.items():
        _write_sheet(wb, sheet_name, rows)
    _write_sheet(wb, "Monthly_PL", monthly_pl)
    _write_sheet(wb, "MIS_Cashflow", cashflow)

    safe_company = re.sub(r"[^A-Za-z0-9]+", "_", company).strip("_")
    stamp = datetime.now().strftime("%Y%m%d%H%M%S")
    filename = f"MIS_{safe_company}_{from_date}_{to_date}_{stamp}.xlsx"
    out_path = EXPORTS_DIR / filename
    wb.save(out_path)
    return out_path