#!/usr/bin/env python3
"""
Tally -> Excel MIS Report  (Generic / Company-agnostic version)
=================================================================

Unlike the old script (which filtered vouchers by a hardcoded party-name
prefix like "VQ SEG26" and classified expenses with company-specific
regexes), this version pulls Tally's own Group + Ledger structure and
classifies every ledger by walking up its group hierarchy to one of
Tally's six *reserved* primary P&L groups:

    Sales Accounts, Purchase Accounts, Direct Incomes, Direct Expenses,
    Indirect Incomes, Indirect Expenses

Those six names are fixed in every Tally company (you cannot rename or
remove them), so this works for any company's chart of accounts without
per-company tuning -- no ledger-naming-convention regex required.

Opening/Closing Stock are not ledgers in Tally (they're inventory
valuation figures), so those two numbers are read from Tally's native
Profit & Loss export instead.

GST is read directly from ledger balances under "Duties & Taxes" whose
names match Output/Input CGST/SGST/IGST patterns -- not back-calculated
at an assumed rate.

Requirements:
    pip install requests xmltodict openpyxl

Run:
    python tally_mis_report_generic.py --company "My Company" \
        --from-date 20260401 --to-date 20270331

Enable Tally's XML server first: F1 > Configuration in Tally, and note
the port (default http://127.0.0.1:9000).
"""

from __future__ import annotations

import argparse
import re
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import requests

try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    openpyxl = None


# --------------------------------------------------------------------------
# Config
# --------------------------------------------------------------------------

@dataclass
class Config:
    tally_url: str = "http://127.0.0.1:9000"
    company: str = ""
    from_date: str = "20260401"   # YYYYMMDD
    to_date: str = "20270331"
    output_file: str = "MIS_Report.xlsx"


# --------------------------------------------------------------------------
# TDL request builders (verified against real Tally output in Postman)
# --------------------------------------------------------------------------

_GROUPS_TDL = """<ENVELOPE>
 <HEADER><VERSION>1</VERSION><TALLYREQUEST>Export</TALLYREQUEST><TYPE>Data</TYPE><ID>GroupList</ID></HEADER>
 <BODY><DESC>
  <STATICVARIABLES>
   <SVCURRENTCOMPANY>{company}</SVCURRENTCOMPANY>
   <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
  </STATICVARIABLES>
  <TDL><TDLMESSAGE>
   <REPORT NAME="GroupList"><FORMS>GroupForm</FORMS></REPORT>
   <FORM NAME="GroupForm"><TOPPARTS>GroupPart</TOPPARTS></FORM>
   <PART NAME="GroupPart">
    <LINES>GroupLine</LINES>
    <REPEAT>GroupLine : GroupColl</REPEAT>
    <SCROLLED>Vertical</SCROLLED>
   </PART>
   <LINE NAME="GroupLine"><FIELDS>FldGName,FldParent,FldIsDeemedPositive</FIELDS></LINE>
   <FIELD NAME="FldGName"><SET>$Name</SET></FIELD>
   <FIELD NAME="FldParent"><SET>$Parent</SET></FIELD>
   <FIELD NAME="FldIsDeemedPositive"><SET>$IsDeemedPositive</SET></FIELD>
   <COLLECTION NAME="GroupColl" ISMODIFY="No">
    <TYPE>Group</TYPE>
    <FETCH>NAME,PARENT,ISDEEMEDPOSITIVE</FETCH>
   </COLLECTION>
  </TDLMESSAGE></TDL>
 </DESC></BODY>
</ENVELOPE>"""
# The custom Report called GroupList.
# So:
# GroupList Report → GroupForm → GroupPart
# The Part controls repetition
# <REPEAT>GroupLine : GroupColl</REPEAT>
# For every record returned by GroupColl, create one GroupLine.
# So if Tally has 30 groups, GroupLine is generated 30 times.
#DeemedPositive is this group's balance normally treated as positive on the debit side?

_LEDGERS_TDL = """<ENVELOPE>
 <HEADER><VERSION>1</VERSION><TALLYREQUEST>Export</TALLYREQUEST><TYPE>Data</TYPE><ID>LedgerList</ID></HEADER>
 <BODY><DESC>
  <STATICVARIABLES>
   <SVCURRENTCOMPANY>{company}</SVCURRENTCOMPANY>
   <SVFROMDATE>{from_date}</SVFROMDATE>
   <SVTODATE>{to_date}</SVTODATE>
   <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
  </STATICVARIABLES>
  <TDL><TDLMESSAGE>
   <REPORT NAME="LedgerList"><FORMS>LedgerForm</FORMS></REPORT>
   <FORM NAME="LedgerForm"><TOPPARTS>LedgerPart</TOPPARTS></FORM>
   <PART NAME="LedgerPart">
    <LINES>LedgerLine</LINES>
    <REPEAT>LedgerLine : LedgerColl</REPEAT>
    <SCROLLED>Vertical</SCROLLED>
   </PART>
   <LINE NAME="LedgerLine"><FIELDS>FldLName,FldLParent,FldOpening,FldClosing</FIELDS></LINE>
   <FIELD NAME="FldLName"><SET>$Name</SET></FIELD>
   <FIELD NAME="FldLParent"><SET>$Parent</SET></FIELD>
   <FIELD NAME="FldOpening"><SET>$OpeningBalance</SET></FIELD>
   <FIELD NAME="FldClosing"><SET>$ClosingBalance</SET></FIELD>
   <COLLECTION NAME="LedgerColl" ISMODIFY="No">
    <TYPE>Ledger</TYPE>
    <FETCH>NAME,PARENT,OPENINGBALANCE,CLOSINGBALANCE</FETCH>
   </COLLECTION>
  </TDLMESSAGE></TDL>
 </DESC></BODY>
</ENVELOPE>"""

_PNL_TDL = """<ENVELOPE>
 <HEADER><VERSION>1</VERSION><TALLYREQUEST>Export</TALLYREQUEST><TYPE>Data</TYPE><ID>Profit and Loss</ID></HEADER>
 <BODY><DESC>
  <STATICVARIABLES>
   <SVCURRENTCOMPANY>{company}</SVCURRENTCOMPANY>
   <SVFROMDATE>{from_date}</SVFROMDATE>
   <SVTODATE>{to_date}</SVTODATE>
   <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
   <SVISSIMPLEPROFITLOSS>No</SVISSIMPLEPROFITLOSS>
  </STATICVARIABLES>
 </DESC></BODY>
</ENVELOPE>"""


def build_groups_request(cfg: Config) -> str:

    return _GROUPS_TDL.format(company=cfg.company)
# It Creates the XML/TDL request used to fetch all Groups from Tally.
# Inserts the selected company name into the request.
# The response is later used to understand the Group → Parent Group hierarchy.


def build_ledgers_request(cfg: Config) -> str:
    return _LEDGERS_TDL.format(company=cfg.company, from_date=cfg.from_date, to_date=cfg.to_date)
# Creates the XML request to fetch all Ledgers from Tally.
# Includes the company, from-date and to-date.
# Retrieves ledger name, parent group, opening balance and closing balance.

def build_pnl_request(cfg: Config) -> str:
    return _PNL_TDL.format(company=cfg.company, from_date=cfg.from_date, to_date=cfg.to_date)
# Creates the XML request for Tally's native Profit & Loss report.
# Uses the selected company and reporting period.
# Mainly used to obtain Opening Stock, Closing Stock and Tally's own P&L figures.


# Stock Items: OPENINGBALANCE and CLOSING BALANCE , no from Date variables and
# any date-scoped quantity/value field triggers Tally's costing engine and
# can hang the whole application on a company with heavy stock movement --
# verified live. Do not add SVFROMDATE here.
_STOCK_ITEMS_TDL = """<ENVELOPE>
 <HEADER><VERSION>1</VERSION><TALLYREQUEST>Export</TALLYREQUEST><TYPE>Collection</TYPE><ID>StockItems</ID></HEADER>
 <BODY><DESC>
  <STATICVARIABLES>
   <SVCURRENTCOMPANY>{company}</SVCURRENTCOMPANY>
   <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
  </STATICVARIABLES>
  <TDL><TDLMESSAGE>
   <COLLECTION NAME="StockItems" ISINITIALIZE="Yes">
    <TYPE>Stock Item</TYPE>
    <FETCH>NAME, PARENT, OPENINGBALANCE, CLOSINGBALANCE</FETCH>
   </COLLECTION>
  </TDLMESSAGE></TDL>
 </DESC></BODY>
</ENVELOPE>"""


def build_stock_items_request(cfg: Config) -> str:
    return _STOCK_ITEMS_TDL.format(company=cfg.company)
# Creates a TDL collection request to retrieve Stock Items from Tally.
# It fetches the item name, stock group, opening balance, and closing balance fields.
# The code intentionally avoids date-specific variables because stock-related requests may cause performance problems.

def build_ledgers_request_for_range(cfg: Config, from_date: str, to_date: str) -> str:
    """Same proven-fast Ledger List request as build_ledgers_request, but
    for an arbitrary date window -- used to build the month-wise P&L by
    calling this once per month rather than inventing a new, untested
    request type."""
    return _LEDGERS_TDL.format(company=cfg.company, from_date=from_date, to_date=to_date)
# Creates the same Ledger List request as build_ledgers_request(), but for a custom date range.
# It is mainly used when generating month-wise MIS data.
# This allows the program to fetch ledger information separately for each month.

def month_windows(from_date: str, to_date: str) -> list[tuple[str, str, str]]:
    """Split a YYYYMMDD..YYYYMMDD range into calendar-month (label, from, to)
    windows, e.g. ('Apr-26', '20260401', '20260430')."""
    import calendar
    from datetime import date

    start = date(int(from_date[:4]), int(from_date[4:6]), int(from_date[6:8]))
    end = date(int(to_date[:4]), int(to_date[4:6]), int(to_date[6:8]))

    windows = []
    y, m = start.year, start.month
    while (y, m) <= (end.year, end.month):
        first_day = date(y, m, 1)
        last_day_num = calendar.monthrange(y, m)[1]
        last_day = date(y, m, last_day_num)
        win_start = max(first_day, start)
        win_end = min(last_day, end)
        label = f"{calendar.month_abbr[m]}-{str(y)[2:]}"
        windows.append((label, win_start.strftime("%Y%m%d"), win_end.strftime("%Y%m%d")))
        if m == 12:
            y, m = y + 1, 1
        else:
            m += 1
    return windows
# Splits the overall reporting period into individual calendar-month ranges.
# For example, 20260401–20270331 becomes Apr-26 through Mar-27.
# It returns tuples containing the month label, start date, and end date.

# --------------------------------------------------------------------------
# HTTP + light XML parsing (avoids xmltodict; Tally's flat-repeat XML is
# just a sequence of sibling tags, easiest to walk with ElementTree directly)
# --------------------------------------------------------------------------

_INVALID_XML_CHARREF_RE = re.compile(r"&#(\d+);")


def _strip_invalid_charrefs(text: str) -> str:
    """Tally's XML server emits numeric character references for control
    codes (observed: &#4; right before group names like 'Primary') that
    are illegal in XML 1.0 and make every strict parser (ElementTree
    included) throw ParseError. Drop any numeric char ref that points at
    a C0 control code outside tab/CR/LF -- these carry no real data."""
    def _sub(m: re.Match) -> str:
        code = int(m.group(1))
        if code in (9, 10, 13) or code >= 32:
            return m.group(0)  # legal, keep as-is
        return ""  # illegal control char reference, drop it
    return _INVALID_XML_CHARREF_RE.sub(_sub, text)
# Cleans invalid XML numeric character references returned by Tally.
# For example, illegal control characters such as &#4; are removed before XML parsing.
# This prevents Python's ElementTree parser from throwing XML ParseError exceptions.

def fetch_xml(xml_body: str, cfg: Config) -> ET.Element:
    resp = requests.post(
        cfg.tally_url,
        data=xml_body.encode("utf-8"),
        headers={"Content-Type": "text/xml"},
        timeout=120,
    )
    resp.raise_for_status()
    text = resp.text
    idx = text.find("<ENVELOPE")
    if idx > 0:
        text = text[idx:]
    text = _strip_invalid_charrefs(text)
    return ET.fromstring(text)
# Sends an XML/TDL request to the Tally HTTP server using requests.post().
# It validates the HTTP response, removes invalid XML characters, and parses the response with ElementTree.
# The function returns the parsed XML root element.

def parse_amt(v: str | None) -> float:
    """Tally amounts come formatted like '40,30,880.00' or blank/None for zero."""
    if v is None:
        return 0.0
    v = v.strip().replace(",", "")
    if not v:
        return 0.0
    try:
        return abs(float(v))
    except ValueError:
        return 0.0
# Converts Tally-formatted monetary values into Python float values.
# It removes commas, handles blank/None values as zero, and converts negative values to absolute values.
# For example, "40,30,880.00" becomes 4030880.0.


def parse_groups(root: ET.Element) -> list[dict]:
    groups = []
    names = [e.text for e in root.findall("FLDGNAME")]
    parents = [e.text for e in root.findall("FLDPARENT")]
    for n, p in zip(names, parents):
        groups.append({"name": (n or "").strip(), "parent": (p or "").strip()})
    return groups
# Extracts group names and parent groups from Tally's XML response.
# It reads FLDGNAME and FLDPARENT elements and combines them into dictionaries.
# The output is a list containing each group's name and its parent.

def parse_ledgers(root: ET.Element) -> list[dict]:
    ledgers = []
    names = [e.text for e in root.findall("FLDLNAME")]
    parents = [e.text for e in root.findall("FLDLPARENT")]
    closings = [e.text for e in root.findall("FLDCLOSING")]
    for n, p, c in zip(names, parents, closings):
        ledgers.append({
            "name": (n or "").strip(),
            "parent": (p or "").strip(),
            "closing": parse_amt(c),
        })
    return ledgers
# Extracts ledger name, parent group, and closing balance from the XML response.
# Each ledger is converted into a dictionary containing name, parent, and closing.
# This normalized structure is used throughout the classification and reporting logic.

def parse_pnl_stock(root: ET.Element) -> dict[str, float]:
    """Pull the full set of top-level figures from the native P&L export:
    Opening/Closing Stock (which aren't ledgers at all), plus Sales,
    Purchases, and Indirect Expenses -- used as Tally's own ground truth
    for the Checks sheet."""
    names = [ (e.text or "").strip() for e in root.findall(".//DSPACCNAME/DSPDISPNAME") ]
    amts = root.findall(".//PLAMT")
    result = {
        "opening_stock": 0.0, "closing_stock": 0.0,
        "sales": 0.0, "purchases": 0.0, "indirect_expenses": 0.0,
    }
    for name, amt_node in zip(names, amts):
        sub = amt_node.find("PLSUBAMT")
        main = amt_node.find("BSMAINAMT")
        val = parse_amt(sub.text if sub is not None else None) or parse_amt(main.text if main is not None else None)
        low = name.lower()
        if "opening stock" in low:
            result["opening_stock"] = val
        elif "closing stock" in low:
            result["closing_stock"] = val
        elif low.startswith("sales accounts"):
            result["sales"] = val
        elif "purchase accounts" in low:
            result["purchases"] = val
        elif low.startswith("indirect expenses"):
            result["indirect_expenses"] = val
    result["net_profit"] = round(
        result["sales"] - (result["opening_stock"] + result["purchases"] - result["closing_stock"])
        - result["indirect_expenses"], 2
    )
    return result


_QTY_RE = re.compile(r"[-+]?\d*\.?\d+")
# Reads important figures from Tally's native Profit & Loss XML output.
# It identifies Opening Stock, Closing Stock, Sales, Purchases, and Indirect Expenses by their displayed names.
# It also calculates a native-style Net Profit for comparison with the script's calculation.

def parse_qty(v: str | None) -> float:
    """Tally quantities come as '240 nos' / '515 Nos' -- strip the unit."""
    if not v:
        return 0.0
    m = _QTY_RE.search(v)
    return float(m.group()) if m else 0.0
# Extracts the numeric quantity from Tally quantity strings such as "240 nos" or "515 Nos".
# It uses a regular expression to ignore the unit text.
# The resulting quantity is returned as a floating-point number.

def parse_stock_items(root: ET.Element) -> list[dict]:
    """Parse the Collection-style Stock Item export (nested <STOCKITEM> tags,
    not the flat repeated-field style used by Requests 1-3)."""
    items = []
    for node in root.findall(".//STOCKITEM"):
        name = node.get("NAME", "").strip()
        parent_el = node.find("PARENT")
        parent = (parent_el.text or "").strip() if parent_el is not None else ""
        opening_el = node.find("OPENINGBALANCE")
        opening_qty = parse_qty(opening_el.text if opening_el is not None else None)
        if name:
            items.append({"name": name, "parent": parent, "opening_qty": opening_qty})
    return items
# Parses the nested STOCKITEM elements returned by the Stock Item collection.
# For each item, it extracts the name, parent Stock Group, and Opening Quantity.
# The resulting list is later used to create the Inventory worksheet.


# --------------------------------------------------------------------------
# Generic classification: walk each ledger's group up to one of Tally's
# six reserved P&L primary groups. These names are fixed by Tally itself
# in every company -- this is what makes the classification generic.
# --------------------------------------------------------------------------

# Defines the six Tally primary P&L groups and maps them to internal MIS categories.
# This is the key component that makes the classification company-independent.
# The program can therefore follow custom subgroups without requiring company-specific ledger regex rules.

_PL_PRIMARY_GROUPS = {
    "Sales Accounts": "Sales",
    "Purchase Accounts": "Purchase",
    "Direct Incomes": "Direct Income",
    "Direct Expenses": "Direct Expense",
    "Indirect Incomes": "Indirect Income",
    "Indirect Expenses": "Indirect Expense",
}

# Contains regular expressions for identifying Output/Input CGST, SGST, and IGST ledgers.
# These patterns are used only for GST identification under Duties & Taxes.
# GST is therefore obtained from ledger balances rather than calculated using an assumed tax percentage.

_GST_LEDGER_PATTERNS = {
    "output_cgst": re.compile(r"output.*\bcgst\b", re.IGNORECASE),
    "output_sgst": re.compile(r"output.*\bsgst\b", re.IGNORECASE),
    "output_igst": re.compile(r"output.*\bigst\b", re.IGNORECASE),
    "input_cgst":  re.compile(r"input.*\bcgst\b", re.IGNORECASE),
    "input_sgst":  re.compile(r"input.*\bsgst\b", re.IGNORECASE),
    "input_igst":  re.compile(r"input.*\bigst\b", re.IGNORECASE),
}


def build_parent_map(groups: list[dict]) -> dict[str, str]:
    return {g["name"]: g["parent"] for g in groups}
# Converts the list of group dictionaries into a lookup dictionary.
# The dictionary structure is essentially {group_name: parent_group}.
# This makes it fast to move upward through Tally's group hierarchy.

def resolve_bucket(group_name: str, parent_map: dict[str, str]) -> str | None:
    """Walk up the group hierarchy until we hit one of the 6 reserved P&L
    primary groups, or run off the top (-> None, i.e. Balance Sheet item)."""
    seen = set()
    current = group_name
    while current and current not in seen:
        if current in _PL_PRIMARY_GROUPS:
            return _PL_PRIMARY_GROUPS[current]
        seen.add(current)
        current = parent_map.get(current)
    return None
# Walks upward through a ledger's group hierarchy until it finds one of Tally's six primary P&L groups.
# Those groups are Sales, Purchase, Direct Income, Direct Expense, Indirect Income, and Indirect Expense.
# If no P&L group is found, it returns None, treating the ledger as a non-P&L/Balance Sheet item.


# Best-effort segment/business-line detector: many companies prefix ledger
# names with a segment tag before " - " (e.g. "Chip Manufacturing - Factory
# Rent" or "VQ SEG26 - CHIP - Expense - ..."). This is a *heuristic bonus*,
# not required for the core report -- if a company doesn't use this
# convention, the Segment_P&L sheet is simply skipped.
def guess_segment(ledger_name: str) -> str | None:
    parts = [p.strip() for p in ledger_name.split(" - ")]
    if len(parts) < 2:
        return None
    # Skip a leading generic code token like "VQ SEG26" and use the next part
    first = parts[0]
    if re.fullmatch(r"[A-Z0-9 ]{2,15}", first) and len(parts) > 2:
        return parts[1]
    return first
# Attempts to identify a business segment from a ledger name using " - " separators.
# For example, "Chip Manufacturing - Factory Rent" may produce "Chip Manufacturing".
# This is only a heuristic and is not required for the main P&L classification.


# --------------------------------------------------------------------------
# Report builder
# --------------------------------------------------------------------------

def classify_ledgers(ledgers: list[dict], parent_map: dict[str, str]) -> tuple[dict[str, list[dict]], dict[str, float]]:
    """Shared classification used by both the annual report and the
    per-month cash-flow sheet: buckets P&L ledgers by the 6 reserved
    primary groups, and separately tallies GST duty ledgers found
    anywhere under Duties & Taxes."""
    buckets: dict[str, list[dict]] = {v: [] for v in _PL_PRIMARY_GROUPS.values()}
    gst_totals = {k: 0.0 for k in _GST_LEDGER_PATTERNS}

    for l in ledgers:
        bucket = resolve_bucket(l["parent"], parent_map)
        if bucket:
            buckets[bucket].append({"Ledger": l["name"], "Group": l["parent"], "Amount": l["closing"]})
            continue
        if l["parent"]:
            cur, seen = l["parent"], set()
            under_duties = False
            while cur and cur not in seen:
                if cur == "Duties & Taxes":
                    under_duties = True
                    break
                seen.add(cur)
                cur = parent_map.get(cur)
            if under_duties:
                for key, pattern in _GST_LEDGER_PATTERNS.items():
                    if pattern.search(l["name"]):
                        gst_totals[key] += l["closing"]
                        break
    return buckets, gst_totals
# Classifies every ledger into the appropriate P&L bucket using the Tally group hierarchy.
# It also searches ledgers under Duties & Taxes for Output/Input CGST, SGST, and IGST.
# The function returns both categorized ledger data and GST totals.



# Builds the main MIS report from groups, ledgers, and stock/P&L information.
# It calculates Sales, Purchases, COGS, Gross Profit, Net Profit, GST, and optional Segment P&L.
# It returns these results as dictionaries that later become Excel worksheets.
def build_report(groups: list[dict], ledgers: list[dict], stock: dict[str, float]) -> dict[str, list[dict]]:
    parent_map = build_parent_map(groups)
    buckets, gst_totals = classify_ledgers(ledgers, parent_map)

    segment_totals: dict[str, dict[str, float]] = {}
    for bucket_name, rows in buckets.items():
        for r in rows:
            seg = guess_segment(r["Ledger"])
            if seg:
                segment_totals.setdefault(seg, {v: 0.0 for v in _PL_PRIMARY_GROUPS.values()})
                segment_totals[seg][bucket_name] += r["Amount"]

    for b in buckets:
        buckets[b].sort(key=lambda r: -r["Amount"])

    def total(rows: list[dict]) -> float:
        return round(sum(r["Amount"] for r in rows), 2)

    sales = total(buckets["Sales"])
    direct_income = total(buckets["Direct Income"])
    purchases = total(buckets["Purchase"])
    direct_expense = total(buckets["Direct Expense"])
    indirect_income = total(buckets["Indirect Income"])
    indirect_expense = total(buckets["Indirect Expense"])

    opening_stock = stock.get("opening_stock", 0.0)
    closing_stock = stock.get("closing_stock", 0.0)

    cogs = round(opening_stock + purchases + direct_expense - closing_stock, 2)
    gross_profit = round(sales + direct_income - cogs, 2)
    net_profit = round(gross_profit + indirect_income - indirect_expense, 2)

    net_cgst = round(gst_totals["output_cgst"] - gst_totals["input_cgst"], 2)
    net_sgst = round(gst_totals["output_sgst"] - gst_totals["input_sgst"], 2)
    net_igst = round(gst_totals["output_igst"] - gst_totals["input_igst"], 2)

    pl_summary = [
        {"Particulars": "Sales Accounts", "Amount": sales},
        {"Particulars": "Direct Incomes", "Amount": direct_income},
        {"Particulars": "Opening Stock", "Amount": opening_stock},
        {"Particulars": "Purchase Accounts", "Amount": purchases},
        {"Particulars": "Direct Expenses", "Amount": direct_expense},
        {"Particulars": "Less: Closing Stock", "Amount": -closing_stock},
        {"Particulars": "Cost of Goods Sold", "Amount": cogs},
        {"Particulars": "Gross Profit", "Amount": gross_profit},
        {"Particulars": "Indirect Incomes", "Amount": indirect_income},
        {"Particulars": "Indirect Expenses", "Amount": indirect_expense},
        {"Particulars": "Net Profit", "Amount": net_profit},
    ]

    gst_summary = [
        {"Tax": "CGST", "Output": round(gst_totals["output_cgst"], 2), "Input": round(gst_totals["input_cgst"], 2), "Net Payable": net_cgst},
        {"Tax": "SGST", "Output": round(gst_totals["output_sgst"], 2), "Input": round(gst_totals["input_sgst"], 2), "Net Payable": net_sgst},
        {"Tax": "IGST", "Output": round(gst_totals["output_igst"], 2), "Input": round(gst_totals["input_igst"], 2), "Net Payable": net_igst},
        {"Tax": "TOTAL", "Output": round(sum(v for k, v in gst_totals.items() if k.startswith("output")), 2),
         "Input": round(sum(v for k, v in gst_totals.items() if k.startswith("input")), 2),
         "Net Payable": round(net_cgst + net_sgst + net_igst, 2)},
    ]

    segment_rows = []
    for seg, vals in sorted(segment_totals.items()):
        seg_sales = vals["Sales"] + vals["Direct Income"]
        seg_direct_cost = vals["Purchase"] + vals["Direct Expense"]
        seg_gp = seg_sales - seg_direct_cost
        seg_np = seg_gp + vals["Indirect Income"] - vals["Indirect Expense"]
        segment_rows.append({
            "Segment": seg, "Revenue": round(seg_sales, 2), "Direct Cost": round(seg_direct_cost, 2),
            "Gross Profit": round(seg_gp, 2), "Indirect Expense": round(vals["Indirect Expense"], 2),
            "Net Profit": round(seg_np, 2),
        })

    dashboard = [
        {"Metric": "Revenue (Sales + Direct Income)", "Value": round(sales + direct_income, 2)},
        {"Metric": "Cost of Goods Sold", "Value": cogs},
        {"Metric": "Gross Profit", "Value": gross_profit},
        {"Metric": "Indirect Expenses", "Value": indirect_expense},
        {"Metric": "Net Profit", "Value": net_profit},
        {"Metric": "Net GST Payable", "Value": round(net_cgst + net_sgst + net_igst, 2)},
    ]

    report = {
        "MIS_Dashboard": dashboard,
        "PL_Summary": pl_summary,
        "Data_Sales": buckets["Sales"],
        "Data_Direct_Income": buckets["Direct Income"],
        "Data_Purchases": buckets["Purchase"],
        "Data_Direct_Expenses": buckets["Direct Expense"],
        "Data_Indirect_Income": buckets["Indirect Income"],
        "Data_Indirect_Expenses": buckets["Indirect Expense"],
        "GST_Summary": gst_summary,
    }
    if segment_rows:
        report["Segment_PL"] = segment_rows
    return report


# Calls Tally once for every month in the requested reporting period.
# It uses month_windows() and build_ledgers_request_for_range() to create monthly requests.
# The monthly ledger results are stored and reused by both Monthly P&L and Cashflow reports.
def fetch_monthly_ledgers(cfg: Config) -> list[tuple[str, str, str, list[dict]]]:
    """Fetch the Ledger List once per calendar month in the period. Shared
    by both Monthly_PL and MIS_Cashflow so the 12 extra Tally calls only
    happen once, not twice."""
    result = []
    for label, win_from, win_to in month_windows(cfg.from_date, cfg.to_date):
        month_ledgers = parse_ledgers(fetch_xml(build_ledgers_request_for_range(cfg, win_from, win_to), cfg))
        result.append((label, win_from, win_to, month_ledgers))
    return result

# Creates an accrual-based P&L summary for every month.
# It classifies monthly ledgers and retrieves monthly Opening/Closing Stock from Tally's native P&L.
# It then calculates monthly COGS, Gross Profit, and Net Profit.
def build_monthly_pl(cfg: Config, groups: list[dict], monthly_ledgers: list[tuple[str, str, str, list[dict]]]) -> list[dict]:
    """Month-wise accrual P&L. Opening/Closing stock per month is pulled via
    the proven-fast native P&L export, called once per month (ledger data is
    reused from monthly_ledgers, already fetched)."""
    parent_map = build_parent_map(groups)
    rows = []
    for label, win_from, win_to, month_ledgers in monthly_ledgers:
        buckets, _ = classify_ledgers(month_ledgers, parent_map)
        totals = {k: round(sum(r["Amount"] for r in v), 2) for k, v in buckets.items()}

        month_pnl_xml = fetch_xml(_PNL_TDL.format(company=cfg.company, from_date=win_from, to_date=win_to), cfg)
        month_stock = parse_pnl_stock(month_pnl_xml)

        cogs = round(month_stock["opening_stock"] + totals["Purchase"] + totals["Direct Expense"] - month_stock["closing_stock"], 2)
        gross_profit = round(totals["Sales"] + totals["Direct Income"] - cogs, 2)
        net_profit = round(gross_profit + totals["Indirect Income"] - totals["Indirect Expense"], 2)

        rows.append({
            "Month": label,
            "Sales": totals["Sales"],
            "Purchases": totals["Purchase"],
            "Gross Profit": gross_profit,
            "Indirect Expenses": totals["Indirect Expense"],
            "Net Profit": net_profit,
        })
    return rows


# Keyword heuristic for splitting Indirect Expense ledgers into "People
# Cost" vs "Operating Expenses" for the cash-flow view -- Tally has no
# built-in category for this, so it's a best-effort name match, not a
# guarantee. Every ledger is still shown individually either way, so a
# mis-bucketed ledger is visible and correctable, never hidden.

# Contains keywords such as salary, wage, staff, payroll, and bonus.
# The program uses these keywords to divide Indirect Expenses into People Cost and Operating Expenses.
# This is explicitly a heuristic and can misclassify unusually named ledgers.
_PEOPLE_COST_KEYWORDS = re.compile(
    r"salar|wage|staff|payroll|incentive|welfare|bonus|commission|hr\b|manpower",
    re.IGNORECASE,
)

# Creates a month-wise cashflow-style MIS view from the already downloaded ledger data.
# It calculates revenue, GST, direct costs, people costs, operating expenses, cash inflow, cash outflow, and net cashflow.
# It also generates MoM revenue growth, gross margin %, and net cashflow %.
def build_cashflow_sheet(groups: list[dict], monthly_ledgers: list[tuple[str, str, str, list[dict]]]) -> list[dict]:
    """Monthly cash-flow-style MIS view: Revenue / GST / Direct Variable
    Cost (ledger-level, generic) / People Cost & Operating Expenses (split
    from Indirect Expenses by keyword heuristic) / Cash Inflow-Outflow /
    Net Cashflow / GST payable. Reuses the same monthly ledger fetch as
    Monthly_PL -- no extra Tally calls."""
    parent_map = build_parent_map(groups)
    month_labels = [label for label, *_ in monthly_ledgers]

    per_month = {}
    all_direct_cost_ledgers: set[str] = set()
    all_people_cost_ledgers: set[str] = set()
    all_opex_ledgers: set[str] = set()

    for label, _wf, _wt, month_ledgers in monthly_ledgers:
        buckets, gst_totals = classify_ledgers(month_ledgers, parent_map)

        direct_cost_rows = buckets["Purchase"] + buckets["Direct Expense"]
        people_rows, opex_rows = [], []
        for r in buckets["Indirect Expense"]:
            (people_rows if _PEOPLE_COST_KEYWORDS.search(r["Ledger"]) else opex_rows).append(r)

        revenue = round(sum(r["Amount"] for r in buckets["Sales"] + buckets["Direct Income"]), 2)
        other_income = round(sum(r["Amount"] for r in buckets["Indirect Income"]), 2)
        output_gst = round(gst_totals["output_cgst"] + gst_totals["output_sgst"] + gst_totals["output_igst"], 2)
        input_gst = round(gst_totals["input_cgst"] + gst_totals["input_sgst"] + gst_totals["input_igst"], 2)
        direct_cost_total = round(sum(r["Amount"] for r in direct_cost_rows), 2)
        people_cost_total = round(sum(r["Amount"] for r in people_rows), 2)
        opex_total = round(sum(r["Amount"] for r in opex_rows), 2)

        cash_inflow = round(revenue + output_gst + other_income, 2)
        gross_margin = round(cash_inflow - direct_cost_total, 2)
        cash_outflow = round(direct_cost_total + people_cost_total + opex_total, 2)
        net_cashflow = round(cash_inflow - cash_outflow, 2)

        per_month[label] = {
            "revenue": revenue, "output_gst": output_gst, "other_income": other_income,
            "cash_inflow": cash_inflow, "direct_cost_rows": {r["Ledger"]: r["Amount"] for r in direct_cost_rows},
            "direct_cost_total": direct_cost_total, "gross_margin": gross_margin,
            "people_rows": {r["Ledger"]: r["Amount"] for r in people_rows}, "people_cost_total": people_cost_total,
            "opex_rows": {r["Ledger"]: r["Amount"] for r in opex_rows}, "opex_total": opex_total,
            "cash_outflow": cash_outflow, "net_cashflow": net_cashflow,
            "input_gst": input_gst, "net_gst": round(output_gst - input_gst, 2),
        }
        all_direct_cost_ledgers.update(per_month[label]["direct_cost_rows"])
        all_people_cost_ledgers.update(per_month[label]["people_rows"])
        all_opex_ledgers.update(per_month[label]["opex_rows"])

    def row(label: str, key_fn, total_override=None) -> dict:
        r = {"Particulars": label}
        for m in month_labels:
            r[m] = key_fn(per_month[m])
        r["Total"] = total_override() if total_override else round(sum(r[m] for m in month_labels), 2)
        return r

    def blank() -> dict:
        r = {"Particulars": ""}
        for m in month_labels:
            r[m] = ""
        r["Total"] = ""
        return r

    def ledger_row(name: str, store_key: str) -> dict:
        return row(name, lambda pm: round(pm[store_key].get(name, 0.0), 2))

    rows = [
        row("Revenue", lambda pm: pm["revenue"]),
        row("GST on Sales (Output)", lambda pm: pm["output_gst"]),
        row("Other Operating Income", lambda pm: pm["other_income"]),
        row("Cash Inflow", lambda pm: pm["cash_inflow"]),
    ]
    mom_row = {"Particulars": "MoM Growth % (Revenue)"}
    prev = None
    for m in month_labels:
        rev = per_month[m]["revenue"]
        mom_row[m] = round(100 * (rev - prev) / prev, 2) if prev else 0.0
        prev = rev
    mom_row["Total"] = ""  # MoM growth has no meaningful sum/total
    rows.append(mom_row)
    rows.append(blank())
    rows.append({"Particulars": "DIRECT VARIABLE COST", **{m: "" for m in month_labels}, "Total": ""})
    for name in sorted(all_direct_cost_ledgers):
        rows.append(ledger_row(name, "direct_cost_rows"))
    rows.append(row("Direct Variable Cost Total", lambda pm: pm["direct_cost_total"]))
    rows.append(row("Gross Margin", lambda pm: pm["gross_margin"]))
    total_cash_inflow = round(sum(per_month[m]["cash_inflow"] for m in month_labels), 2)
    total_gross_margin = round(sum(per_month[m]["gross_margin"] for m in month_labels), 2)
    rows.append(row(
        "% Margin", lambda pm: round(100 * pm["gross_margin"] / pm["cash_inflow"], 2) if pm["cash_inflow"] else 0.0,
        total_override=lambda: round(100 * total_gross_margin / total_cash_inflow, 2) if total_cash_inflow else 0.0,
    ))
    rows.append(blank())
    rows.append({"Particulars": "PEOPLE COST (name-matched, see Source_Notes)", **{m: "" for m in month_labels}, "Total": ""})
    for name in sorted(all_people_cost_ledgers):
        rows.append(ledger_row(name, "people_rows"))
    rows.append(row("People Cost Subtotal", lambda pm: pm["people_cost_total"]))
    rows.append(blank())
    rows.append({"Particulars": "OPERATING EXPENSES", **{m: "" for m in month_labels}, "Total": ""})
    for name in sorted(all_opex_ledgers):
        rows.append(ledger_row(name, "opex_rows"))
    rows.append(row("Operating Expenses Subtotal", lambda pm: pm["opex_total"]))
    rows.append(blank())
    rows.append(row("Cash Outflow", lambda pm: pm["cash_outflow"]))
    rows.append(row("Net Cashflow / Operating Profit", lambda pm: pm["net_cashflow"]))
    total_net_cashflow = round(sum(per_month[m]["net_cashflow"] for m in month_labels), 2)
    rows.append(row(
        "Net Cashflow %", lambda pm: round(100 * pm["net_cashflow"] / pm["cash_inflow"], 2) if pm["cash_inflow"] else 0.0,
        total_override=lambda: round(100 * total_net_cashflow / total_cash_inflow, 2) if total_cash_inflow else 0.0,
    ))
    rows.append(blank())
    rows.append(row("GST on Sales (Output)", lambda pm: pm["output_gst"]))
    rows.append(row("Input GST on Purchases", lambda pm: pm["input_gst"]))
    rows.append(row("Net GST Payable", lambda pm: pm["net_gst"]))
    return rows


# Creates the Inventory worksheet using Stock Item master data.
# It displays Opening Quantity and attempts to match Purchase/Sales ledger values using stock-item names.
# Closing Quantity is deliberately left blank because the code considers the available Tally API result unreliable.
def build_inventory_sheet(stock_items: list[dict], ledgers: list[dict]) -> list[dict]:
    """Opening Qty comes straight from the Stock Item master (fast, verified).
    Closing Qty is deliberately left blank -- every Tally-side attempt to get
    it (master ClosingBalance, voucher WALK, native Stock Summary) either hung
    the application or returned a figure that didn't reconcile with known
    totals, so showing a number here would be worse than showing none.
    Purchase/Sales value per item is matched by name against the ledgers
    we already have -- no extra Tally calls needed."""
    rows = []
    for item in stock_items:
        name_l = item["name"].lower()
        purchase_val = sum(l["closing"] for l in ledgers if l["parent"] == "Purchase Accounts" and name_l in l["name"].lower())
        sales_val = sum(l["closing"] for l in ledgers if l["parent"] == "Sales Accounts" and name_l in l["name"].lower())
        rows.append({
            "Stock Item": item["name"],
            "Stock Group": item["parent"],
            "Opening Qty": item["opening_qty"],
            "Closing Qty": None,  # see docstring -- not reliably obtainable, left blank on purpose
            "Purchase Value (matched)": round(purchase_val, 2),
            "Sales Value (matched)": round(sales_val, 2),
        })
    return rows

# Validates the script's calculations against Tally's own native P&L output.
# It compares Sales, Purchases, Indirect Expenses, and Net Profit.
# A difference below ₹1 is marked PASS; otherwise, the result is marked REVIEW.
def build_checks_sheet(report: dict[str, list[dict]], native_pnl: dict[str, float]) -> list[dict]:
    """Self-validation: compare the script's computed P&L figures against
    Tally's own native Profit & Loss export (ground truth)."""
    pl = {r["Particulars"]: r["Amount"] for r in report["PL_Summary"]}
    checks = [
        ("Sales Accounts", pl.get("Sales Accounts", 0.0), native_pnl.get("sales", 0.0)),
        ("Purchase Accounts", pl.get("Purchase Accounts", 0.0), native_pnl.get("purchases", 0.0)),
        ("Indirect Expenses", pl.get("Indirect Expenses", 0.0), native_pnl.get("indirect_expenses", 0.0)),
        ("Net Profit", pl.get("Net Profit", 0.0), native_pnl.get("net_profit", 0.0)),
    ]
    rows = []
    for label, computed, native in checks:
        diff = round(computed - native, 2)
        rows.append({
            "Check": label,
            "Script Value": round(computed, 2),
            "Tally Native P&L": round(native, 2),
            "Difference": diff,
            "Status": "PASS" if abs(diff) < 1.0 else "REVIEW",
        })
    return rows

# Creates documentation describing where each MIS figure comes from.
# It records company, reporting period, generation time, classification methodology, GST source, stock source, and known limitations.
# These records become the Source_Notes Excel worksheet.
def build_source_notes(cfg: Config) -> list[dict]:
    from datetime import datetime
    return [
        {"Field": "Company", "Value": cfg.company},
        {"Field": "Period From", "Value": cfg.from_date},
        {"Field": "Period To", "Value": cfg.to_date},
        {"Field": "Generated At", "Value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        {"Field": "Sales/Purchase/Expense classification", "Value": "Ledger -> Group hierarchy walked to Tally's 6 reserved P&L primary groups"},
        {"Field": "GST figures", "Value": "Read directly from Duties & Taxes ledger closing balances, not back-calculated at a fixed rate"},
        {"Field": "Opening/Closing Stock", "Value": "From Tally's native Profit & Loss export (not a ledger figure)"},
        {"Field": "Inventory Closing Qty", "Value": "Not populated -- unreliable via API on this company without risking a Tally hang; check Stock Summary in Tally directly"},
        {"Field": "Segment_PL", "Value": "Best-effort heuristic based on ' - ' prefix in ledger names; may split one real segment into two labels if naming conventions are inconsistent"},
        {"Field": "MIS_Cashflow People Cost split", "Value": "Best-effort keyword match on ledger name (salary/wage/staff/incentive/etc.) vs Operating Expenses; a differently-named people-cost ledger (e.g. 'Manpower Cost') would fall into Operating Expenses instead -- every ledger is still listed individually so this is checkable"},
        {"Field": "MIS_Cashflow Revenue", "Value": "Shown as booked in Sales/Direct Income ledgers, net of GST (GST on Sales shown as a separate line) -- not assumed gross-of-GST"},
    ]


# --------------------------------------------------------------------------
# Excel writer
# --------------------------------------------------------------------------


# Creates a new Excel workbook using openpyxl.
# Each key in the report dictionary becomes a worksheet, and dictionary fields become column headers.
# It formats headers, adjusts column widths, saves the workbook, and returns the absolute output path.
def write_report_to_excel(report: dict[str, list[dict]], cfg: Config) -> str:
    if openpyxl is None:
        raise RuntimeError("Install openpyxl: pip install openpyxl")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name, rows in report.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        columns = list(rows[0].keys()) if rows else ["Particulars"]
        ws.append(columns)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for r in rows:
            ws.append([r.get(c, "") for c in columns])
        for col_cells in ws.columns:
            width = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(width + 2, 10), 45)
        print(f"  wrote {len(rows)} row(s) to '{sheet_name}'")

    out_path = Path(cfg.output_file).resolve()
    wb.save(out_path)
    return str(out_path)


# --------------------------------------------------------------------------
# Orchestration
# --------------------------------------------------------------------------
# This is the main orchestration function of the application.
# It fetches Groups, Ledgers, native P&L, and Stock Items, then builds all MIS worksheets.
# If monthly reporting is enabled, it additionally creates Monthly P&L and MIS Cashflow before exporting everything to Excel.
def run(cfg: Config, include_monthly: bool = True) -> dict[str, list[dict]]:
    print(f"Fetching Groups for '{cfg.company}'...")
    groups = parse_groups(fetch_xml(build_groups_request(cfg), cfg))
    print(f"  {len(groups)} group(s)")

    print(f"Fetching Ledgers for '{cfg.company}' ({cfg.from_date} - {cfg.to_date})...")
    ledgers = parse_ledgers(fetch_xml(build_ledgers_request(cfg), cfg))
    print(f"  {len(ledgers)} ledger(s)")

    print("Fetching native Profit & Loss (for Opening/Closing Stock + ground-truth check)...")
    native_pnl = parse_pnl_stock(fetch_xml(build_pnl_request(cfg), cfg))
    print(f"  Opening Stock={native_pnl['opening_stock']}, Closing Stock={native_pnl['closing_stock']}")

    print("Fetching Stock Items (Opening Qty only -- see Source_Notes for why Closing Qty is omitted)...")
    stock_items = parse_stock_items(fetch_xml(build_stock_items_request(cfg), cfg))
    print(f"  {len(stock_items)} stock item(s)")

    print("Building MIS report...")
    report = build_report(groups, ledgers, native_pnl)

    report["Inventory"] = build_inventory_sheet(stock_items, ledgers)
    report["Checks"] = build_checks_sheet(report, native_pnl)
    report["Source_Notes"] = build_source_notes(cfg)

    if include_monthly:
        windows = month_windows(cfg.from_date, cfg.to_date)
        print(f"Fetching monthly ledger data ({len(windows)} month(s), reusing the proven-fast Ledger List request)...")
        monthly_ledgers = fetch_monthly_ledgers(cfg)

        print("Building Monthly P&L...")
        report["Monthly_PL"] = build_monthly_pl(cfg, groups, monthly_ledgers)

        print("Building MIS Cashflow...")
        report["MIS_Cashflow"] = build_cashflow_sheet(groups, monthly_ledgers)

    print("Writing Excel workbook...")
    path = write_report_to_excel(report, cfg)
    print(f"Done. Saved to {path}")
    return report

# This is the command-line entry point of the program.
# It defines arguments such as Tally URL, company, from-date, to-date, output filename, and --no-monthly.
# It creates the Config object and calls run() with the user's selected options.
def main() -> None:
    parser = argparse.ArgumentParser(description="Generic Tally -> Excel MIS report")
    parser.add_argument("--tally-url", default="http://127.0.0.1:9000")
    parser.add_argument("--company", required=True)
    parser.add_argument("--from-date", default="20260401", help="YYYYMMDD")
    parser.add_argument("--to-date", default="20270331", help="YYYYMMDD")
    parser.add_argument("--output-file", default="MIS_Report.xlsx")
    parser.add_argument("--no-monthly", action="store_true",
                         help="Skip Monthly_PL (saves ~2 Tally calls per month in the period)")
    args = parser.parse_args()

    cfg = Config(
        tally_url=args.tally_url, company=args.company,
        from_date=args.from_date, to_date=args.to_date,
        output_file=args.output_file,
    )
    run(cfg, include_monthly=not args.no_monthly)


if __name__ == "__main__":
    main()
